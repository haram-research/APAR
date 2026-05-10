import openpyxl, re

wb = openpyxl.load_workbook(
    '/Users/haram/Desktop/하람/대학원/APAR/주관식채점_기초자료_(2026_중간).xlsx',
    read_only=True, data_only=True
)
ws = wb.active

problems = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if len(row) < 6:
        continue
    prob_num = str(row[2])
    answer_key = row[3]
    student_ans = row[4]
    auto_score = str(row[5]) if row[5] is not None else ''
    manual_score = row[7] if len(row) > 7 else None

    if prob_num not in problems:
        problems[prob_num] = {
            'answer_key': str(answer_key)[:500] if answer_key else '',
            'students': []
        }
    problems[prob_num]['students'].append({
        'ans': str(student_ans) if student_ans is not None else '',
        'auto': auto_score,
        'manual_score': manual_score
    })

def show_problem(pnum_str, label=''):
    p = problems.get(pnum_str)
    if not p:
        print('문제 %s 없음' % pnum_str)
        return
    print('=== 문제 %s %s ===' % (pnum_str, label))
    print('D열 정답: %s' % p['answer_key'][:250])
    total = len(p['students'])
    correct = [s for s in p['students'] if s['auto'] != '0' and s['auto'] != '']
    wrong = [s for s in p['students'] if s['auto'] == '0']
    print('총: %d명 | 자동정답: %d명(%.1f%%) | 자동오답: %d명(%.1f%%)' % (
        total, len(correct), len(correct)/total*100, len(wrong), len(wrong)/total*100
    ))
    print('  [정답 샘플]:')
    for s in correct[:5]:
        print('    [%s] -> %s' % (s['ans'][:100], s['auto']))
    print('  [오답 샘플]:')
    for s in wrong[:10]:
        print('    [%s] -> %s' % (s['ans'][:100], s['auto']))
    print()

# ---- 유형 1: Expletive infixation (단일 답, 대안 표현 다수) ----
print('########## 유형1: 단일 정답형 (대안 표현 다수) ##########')
show_problem('1', '[manufuck*facturer]')
show_problem('96', '[au-bloody-tomatic]')
show_problem('113', '[dis-fucking-honest, ab-fucking-normal]')
show_problem('103', '[fingerspelling]')
show_problem('107', '[magnet]')
show_problem('117', '[modularity]')
show_problem('120', '[modularity]')

# ---- 유형 2: 문장/구 정답형 ----
print('########## 유형2: 문장/구 정답형 ##########')
show_problem('98', '[Susan is aware of the fact that Bill is]')
show_problem('102', '[genetically determined specialization]')
show_problem('104', '[Paradox Language Acquisition]')

# ---- 유형 3: 단어 리스트형 (순서 무관, 부분점수 필요) ----
print('########## 유형3: 단어 리스트형 ##########')
show_problem('100', '[Jackendoff + 5단어]')
show_problem('108', '[toves wabe Jabberwock ...]')
show_problem('109', '[brillig slithy frumious Long manxome]')
show_problem('110', '[was Did gyre gimble Beware bite catch Shun sought]')

# ---- 유형 4: 번호 선택형 ----
print('########## 유형4: 번호 선택형 ##########')
show_problem('97', '[5 7 9]')
show_problem('111', '[p d]')
show_problem('114', '[1 2 3 6]')
show_problem('115', '[1 3 4 9]')
show_problem('116', '[1 3 4 5 8 9]')

# ---- 스펠링 오류 사례 집중 분석 (문제 109, 108) ----
print('########## 스펠링 오류 집중 분석 ##########')
for pnum_str in ['109', '108']:
    p = problems[pnum_str]
    print('--- 문제 %s 오답 전체 (스펠링 오류 탐지) ---' % pnum_str)
    wrong = [s for s in p['students'] if s['auto'] == '0' and s['ans']]
    for s in wrong[:20]:
        print('  [%s]' % s['ans'][:120])
    print()

# ---- 대소문자 민감도 분석 (문제 109 Long 이슈) ----
print('########## 대소문자 이슈 분석 (문제 109) ##########')
p109 = problems['109']
print('D열 정답:', p109['answer_key'])
all_s = [s for s in p109['students'] if s['ans']][:30]
for s in all_s:
    print('  [%s] -> %s' % (s['ans'][:100], s['auto']))
