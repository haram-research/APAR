"""
test_pregrade.py
─────────────────────────────────────────────────────────────────
APAR preGradeService.js 로직을 Python으로 재현하여
실제 xlsx 데이터에 적용하고 eclass 자동채점과 비교합니다.

목표:
  1. 정규화(normalizeAnswer) + Type A/D 채점이 eclass 정답과 일치하는지 확인
  2. eclass가 0점 처리했으나 우리 로직이 구제할 수 있는 케이스 발굴
  3. eclass 정답인데 우리 로직이 놓치는 케이스 (위험 케이스) 확인
"""

import openpyxl
import re
import unicodedata

XLSX_PATH = '/Users/haram/Desktop/하람/대학원/APAR/주관식채점_기초자료_(2026_중간).xlsx'

# ─────────────────────────────────────────────────────────────────
# 1. normalizeAnswer (JS 로직 동일 구현)
# ─────────────────────────────────────────────────────────────────
def normalize_answer(text):
    if not text or text in ('None', 'null', ''):
        return ''
    text = str(text)

    # 보이지 않는 문자 제거
    text = text.replace('\u200B', '').replace('\uFEFF', '')

    # 특수 공백 → 일반 공백
    for ch in '\u2006\u2002\u2003\u00A0\u3000':
        text = text.replace(ch, ' ')

    # 대시류 → 하이픈
    for ch in '\u2013\u2014\u2212':
        text = text.replace(ch, '-')

    # 곡선 따옴표
    for ch in '\u2018\u2019':
        text = text.replace(ch, "'")
    for ch in '\u201C\u201D':
        text = text.replace(ch, '"')

    # 전각 ASCII → 반각
    result = []
    for c in text:
        cp = ord(c)
        if 0xFF01 <= cp <= 0xFF5E:
            result.append(chr(cp - 0xFEE0))
        else:
            result.append(c)
    text = ''.join(result)

    # 동아시아 구분자
    text = text.replace('\u3001', ',').replace('\uFF0C', ',')
    text = text.replace('（', '(').replace('）', ')')

    # 연속 공백
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def normalize_ci(text):
    return normalize_answer(text).lower()


# ─────────────────────────────────────────────────────────────────
# 2. parseAnswerKey
# ─────────────────────────────────────────────────────────────────
def parse_answer_key(d_col):
    if not d_col or d_col == 'None':
        return []
    matches = re.findall(r'\[([^\[\]]*)\]', str(d_col))
    return [m.strip() for m in matches if m.strip()]


# ─────────────────────────────────────────────────────────────────
# 3. detectQuestionType
# ─────────────────────────────────────────────────────────────────
def detect_question_type(allowed_forms):
    if not allowed_forms:
        return 'UNKNOWN'
    is_all_digits = all(re.fullmatch(r'[\d\s,().]+', f) for f in allowed_forms)
    return 'D' if is_all_digits else 'A'


# ─────────────────────────────────────────────────────────────────
# 4. gradeTypeA
# ─────────────────────────────────────────────────────────────────
def canonicalize(s):
    """괄호, 콤마, 공백을 정규화하여 핵심 토큰만 비교"""
    s = re.sub(r'[()\[\]]', ' ', s)   # 괄호를 공백으로 치환
    s = s.replace(',', ' ')           # 콤마 → 공백
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def strip_all(s):
    """모든 구분자 제거 — 공백/콤마/괄호/하이픈 무시"""
    return re.sub(r'[()\[\]\s,\-]', '', s)


def grade_type_a(student_ans, allowed_forms):
    norm_student = normalize_ci(student_ans)

    # 1차: 완전 일치
    for form in allowed_forms:
        if normalize_ci(form) == norm_student:
            return {'matched': True, 'form': form}

    # 2차: canonicalize 후 비교 (괄호/콤마 공백 무관)
    can_student = canonicalize(norm_student)
    for form in allowed_forms:
        if canonicalize(normalize_ci(form)) == can_student:
            return {'matched': True, 'form': form, 'note': 'canonical'}

    # 3차: 모든 구분자 제거 후 비교 (fingerspelling / finger spelling 등)
    stripped_student = strip_all(norm_student)
    is_short = len(stripped_student) <= 30
    if is_short:
        for form in allowed_forms:
            if strip_all(normalize_ci(form)) == stripped_student:
                return {'matched': True, 'form': form, 'note': 'separator-stripped'}

    return {'matched': False}


# ─────────────────────────────────────────────────────────────────
# 5. gradeTypeD
# ─────────────────────────────────────────────────────────────────
def extract_digit_set(text):
    normalized = normalize_answer(text)
    tokens = re.split(r'[\s,()]+', normalized)
    digits = set()
    for token in tokens:
        num = re.sub(r'\D', '', token)
        if not num:
            continue
        if len(num) == 1:
            # 단자리: 그대로
            digits.add(num)
        else:
            # 다자리: 붙여쓴 번호열 → 각 자리를 개별 번호로 분리
            # "579" → {'5','7','9'}, "1236" → {'1','2','3','6'}
            for ch in num:
                digits.add(ch)
    return sorted(digits)


def grade_type_d(student_ans, allowed_forms):
    student_digits = extract_digit_set(student_ans)
    correct_digits = extract_digit_set(allowed_forms[0])
    is_match = student_digits == correct_digits
    return {'matched': is_match, 'student': student_digits, 'correct': correct_digits}


# ─────────────────────────────────────────────────────────────────
# 6. preGrade (메인)
# ─────────────────────────────────────────────────────────────────
def pre_grade(student_ans, d_col, max_score=1):
    allowed_forms = parse_answer_key(d_col)
    if not allowed_forms:
        return None

    q_type = detect_question_type(allowed_forms)

    if q_type == 'D':
        result = grade_type_d(student_ans, allowed_forms)
        return {
            'score': max_score if result['matched'] else 0,
            'reason': (
                f"[D] 정답: {result['matched']} | 제출={result['student']} 정답={result['correct']}"
            ),
            'skip_llm': True,
            'q_type': 'D',
            'matched': result['matched'],
        }

    if q_type == 'A':
        result = grade_type_a(student_ans, allowed_forms)
        if result['matched']:
            return {
                'score': max_score,
                'reason': f"[A] 일치: \"{result['form']}\"",
                'skip_llm': True,
                'q_type': 'A',
                'matched': True,
            }
        return None  # LLM 위임

    return None


# ─────────────────────────────────────────────────────────────────
# 7. 데이터 로드 및 테스트 실행
# ─────────────────────────────────────────────────────────────────
def main():
    print("xlsx 로딩 중...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active

    # 컬럼: A=Serial, B=학번, C=문제번호, D=정답키, E=학생답안, F=자동점수
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 6:
            continue
        rows.append({
            'prob': str(row[2]),
            'd_col': str(row[3]) if row[3] else '',
            'student': str(row[4]) if row[4] else '',
            'eclass': str(row[5]) if row[5] is not None else '0',
        })

    print(f"총 {len(rows)}행 로드 완료\n")

    # 문제별 통계
    from collections import defaultdict
    stats = defaultdict(lambda: {
        'q_type': '?',
        'total': 0,
        'eclass_correct': 0,   # eclass > 0
        'apar_correct': 0,     # apar matched
        'rescued': 0,          # eclass=0 but apar matched (개선 케이스)
        'false_neg': 0,        # eclass>0 but apar missed
        'false_pos': 0,        # eclass=0 and apar also says 0 (consistent)
        'llm_needed': 0,       # apar returns None (LLM 위임)
        'rescue_examples': [],
        'false_neg_examples': [],
    })

    for r in rows:
        prob = r['prob']
        allowed = parse_answer_key(r['d_col'])
        q_type = detect_question_type(allowed) if allowed else 'UNKNOWN'
        stats[prob]['q_type'] = q_type

        # eclass 정답 여부 (점수 > 0)
        try:
            eclass_score = float(r['eclass'])
        except (ValueError, TypeError):
            eclass_score = 0
        eclass_ok = eclass_score > 0

        stats[prob]['total'] += 1
        if eclass_ok:
            stats[prob]['eclass_correct'] += 1

        # APAR 사전 채점 (max_score=1 — 비율만 비교)
        pg = pre_grade(r['student'], r['d_col'], max_score=1)

        if pg is None:
            stats[prob]['llm_needed'] += 1
            # Type A 불일치 케이스 중 eclass가 정답인 것 = false negative
            if eclass_ok:
                stats[prob]['false_neg'] += 1
                if len(stats[prob]['false_neg_examples']) < 3:
                    stats[prob]['false_neg_examples'].append(
                        f"  eclass=✅ | 학생=[{r['student'][:60]}]"
                    )
        else:
            apar_ok = pg['matched']
            if apar_ok:
                stats[prob]['apar_correct'] += 1

            if not eclass_ok and apar_ok:
                # 구제 케이스: eclass 오답 → APAR 정답
                stats[prob]['rescued'] += 1
                if len(stats[prob]['rescue_examples']) < 3:
                    stats[prob]['rescue_examples'].append(
                        f"  학생=[{r['student'][:60]}] | {pg['reason'][:80]}"
                    )
            elif eclass_ok and not apar_ok:
                stats[prob]['false_neg'] += 1
                if len(stats[prob]['false_neg_examples']) < 3:
                    stats[prob]['false_neg_examples'].append(
                        f"  eclass=✅ | 학생=[{r['student'][:60]}] | {pg['reason'][:60]}"
                    )

    # ─── 결과 출력 ───────────────────────────────────────────
    print("=" * 80)
    print(f"{'문제':>5} {'유형':>4} | {'전체':>5} {'eclass정답':>8} {'APAR정답':>7} "
          f"{'구제↑':>6} {'놓침↓':>6} {'LLM위임':>7}")
    print("-" * 80)

    total_rescued = 0
    total_false_neg = 0
    total_llm = 0
    type_d_probs = []
    type_a_probs = []

    for prob in sorted(stats.keys(), key=lambda x: int(x) if x.isdigit() else 999):
        s = stats[prob]
        q = s['q_type']
        rescued_pct = f"{s['rescued']/s['total']*100:.1f}%" if s['total'] else '-'
        fn_pct = f"{s['false_neg']/s['total']*100:.1f}%" if s['total'] else '-'

        print(f"{prob:>5} {q:>4} | {s['total']:>5} "
              f"{s['eclass_correct']:>5}({s['eclass_correct']/s['total']*100:.0f}%) "
              f"{s['apar_correct']:>5}({s['apar_correct']/s['total']*100:.0f}%) "
              f"{s['rescued']:>3}({rescued_pct}) "
              f"{s['false_neg']:>3}({fn_pct}) "
              f"{s['llm_needed']:>5}")

        total_rescued += s['rescued']
        total_false_neg += s['false_neg']
        total_llm += s['llm_needed']
        if q == 'D':
            type_d_probs.append(prob)
        elif q == 'A':
            type_a_probs.append(prob)

    print("=" * 80)
    print(f"\n총 구제 케이스(eclass 오답→APAR 정답): {total_rescued}건")
    print(f"총 놓침 케이스(eclass 정답→APAR 오답): {total_false_neg}건  ← 위험!")
    print(f"총 LLM 위임 케이스: {total_llm}건")
    print(f"\nType D 문제: {type_d_probs}")
    print(f"Type A 문제: {type_a_probs}")

    # 구제 케이스 상세 출력
    print("\n" + "=" * 80)
    print("▼ 구제 케이스 예시 (eclass=0점, APAR=정답 처리)")
    print("=" * 80)
    for prob in sorted(stats.keys(), key=lambda x: int(x) if x.isdigit() else 999):
        s = stats[prob]
        if s['rescue_examples']:
            print(f"\n[문제 {prob}] ({s['rescued']}건)")
            for ex in s['rescue_examples']:
                print(ex)

    # 위험 케이스 (false negative) 상세
    print("\n" + "=" * 80)
    print("▼ 놓침 케이스 예시 (eclass=정답, APAR=오답 — 검토 필요)")
    print("=" * 80)
    for prob in sorted(stats.keys(), key=lambda x: int(x) if x.isdigit() else 999):
        s = stats[prob]
        if s['false_neg_examples']:
            print(f"\n[문제 {prob}] ({s['false_neg']}건)")
            for ex in s['false_neg_examples']:
                print(ex)

    # 유니코드 구제 확인 (특수문자 포함 → 정규화 후 정답 처리)
    print("\n" + "=" * 80)
    print("▼ 유니코드 정규화 구제 케이스 (특수문자 포함 답변)")
    print("=" * 80)
    unicode_rescued = []
    for r in rows:
        has_special = any(ord(c) > 127 and ord(c) < 0xAC00 for c in r['student'])
        if not has_special:
            continue
        try:
            eclass_score = float(r['eclass'])
        except:
            eclass_score = 0
        pg = pre_grade(r['student'], r['d_col'], 1)
        if pg and pg['matched'] and eclass_score == 0:
            unicode_rescued.append(
                f"문제{r['prob']}: [{r['student'][:70]}] → {pg['reason'][:70]}"
            )

    if unicode_rescued:
        for u in unicode_rescued[:20]:
            print(' ', u)
    else:
        print("  (유니코드 정규화로 구제된 케이스 없음 — 정답 목록과 학생 답안 둘 다 특수문자인 경우는 없음)")


if __name__ == '__main__':
    main()
